from __future__ import annotations

import csv
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

FIELDS = [
    "author",
    "published_at",
    "like_count",
    "text",
    "reply_count",
    "comment_id",
    "parent_id",
    "video_id",
]


def build_filename(video_id: str, ext: str) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    return f"comments_{video_id}_{ts}.{ext}"


def export_csv(comments: List[dict], export_dir: str, video_id: str, fields: List[str]) -> str:
    os.makedirs(export_dir, exist_ok=True)
    filename = build_filename(video_id, "csv")
    path = os.path.join(export_dir, filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for c in comments:
            writer.writerow({k: c.get(k, "") for k in fields})

    return path


def export_xlsx(comments: List[dict], export_dir: str, video_id: str, fields: List[str]) -> str:
    os.makedirs(export_dir, exist_ok=True)
    filename = build_filename(video_id, "xlsx")
    path = os.path.join(export_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    ws.append(fields)
    for c in comments:
        ws.append([c.get(k, "") for k in fields])

    for col_idx, col_name in enumerate(fields, start=1):
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    wb.save(path)
    return path


def export_json(comments: List[dict], export_dir: str, video_id: str, fields: List[str]) -> str:
    os.makedirs(export_dir, exist_ok=True)
    filename = build_filename(video_id, "json")
    path = os.path.join(export_dir, filename)
    payload = [{k: c.get(k, "") for k in fields} for c in comments]
    with open(path, "w", encoding="utf-8") as f:
        import json

        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path
